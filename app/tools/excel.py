from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from app.tools import files


def add_excel(row, path='./xlsx/result.xlsx', sheet_name='Main', title: None | tuple = None) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb:
        if wb.worksheets[0].title != 'Sheet':
            wb.create_sheet(sheet_name)
            ws: Worksheet = wb[sheet_name]
        else:
            ws: Worksheet = wb.worksheets[0]
            ws.title = sheet_name
        if title:
            ws.append(title)
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'
    ws: Worksheet = wb[sheet_name]
    ws.append(row)
    wb.save(path)
    wb.close()


def new_excel(path='./xlsx/result.xlsx', delete: bool = True):
    if delete:
        files.delete_path('/'.join([part for part in path.split('/')[:2]]))
    files.create_dir('/'.join([part for part in path.split('/')[:2]]))
    wb = Workbook()
    wb.save(filename=path)
